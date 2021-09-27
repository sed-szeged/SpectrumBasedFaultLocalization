import subprocess
import testFiles
import sys
import os
import json
import pytest
import openpyxl
import os.path

def get_tests_results(path):

    test_folder, tests_folder, test_files, tests_files = testFiles.get_TestFiles(path)

    pytest_args = ["-v"]
    for file in tests_files:
        pytest_args.append(path + os.path.sep + tests_folder + os.path.sep + file)
    for file in test_files:
        pytest_args.append(path + os.path.sep + test_folder + os.path.sep + file)


    pytest_args.append("--excelreport=" + path + os.path.sep + "report.xlsx")

    print(pytest_args)

    pytest.main(pytest_args)
    results_dict = {}
    report_file = path + os.path.sep + "report.xlsx"
    
    if not os.path.exists(report_file):
        sys.exit(7)

    wb_obj = openpyxl.load_workbook(report_file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row - 1

    for i in range(m_row):
        t_s = sheet_obj.cell(row = i + 2, column = 1).value
        t_n = sheet_obj.cell(row = i + 2, column = 2).value
        t_r = sheet_obj.cell(row = i + 2, column = 4).value
        temp1 = sheet_obj.cell(row = i + 2, column = 8).value
        temp2 = temp1.replace(os.path.sep, ".")
        t_f_n = temp2.replace(".py", "")
        file = temp1.replace(".py", "")

        if t_f_n == t_s:
            c_t_n = t_f_n + "." + t_n
        else:
            c_t_n = t_f_n + "." + t_s + "." + t_n
        #c_t_n = t_f_n + "." + t_s + "." + t_n


        if t_r == "PASSED" or t_r == "FAILED":
            results_dict[c_t_n] = t_r

    if results_dict:
        print("Thisisresultsdict", results_dict)
        return results_dict
    else:
        sys.exit(3)
